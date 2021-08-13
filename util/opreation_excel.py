from openpyxl import load_workbook


class OperationExcel:
    def __init__(self, file_name=None):
        self.table = None
        if file_name:
            self.file_name = file_name
        else:
            self.file_name = '../case_excel/case.xlsx'
        self.workbook = load_workbook(self.file_name)

    def get_table(self, sheet_name=None):
        self.table = self.workbook.active
        if sheet_name:
            self.table = self.workbook[sheet_name]

    def get_rows(self):
        rows = self.table.max_row
        return rows

    def get_cols(self):
        cols = self.table.max_column
        return cols

    def get_cell_value(self, x, y):
        cell_value = self.table.cell(x, y).value
        return cell_value

    def write_execel(self, row, col, value, font=None):
        self.table.cell(row, col, value).font = font
        self.workbook.save(self.file_name)


if __name__ == '__main__':
    excel = OperationExcel('../case_excel/case.xlsx')
    excel.get_table(sheet_name='case1')
    excel.write_execel(15, 15, '返回成功！')
    print(excel.get_cell_value(1, 1))
